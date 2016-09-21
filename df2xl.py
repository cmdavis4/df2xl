import numpy as np
import pandas as pd
import datetime as dt
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from xlsx2csv import Xlsx2csv
from collections import OrderedDict
import openpyxl
import csv
import os

from .Utils import assertType, isIntegerLike, isNumerical, isDatetimeLike

#TODO: Assertions on types of arguments


def asRow(s):
    assert type(s) == pd.Series
    return pd.DataFrame(s).transpose()

def asColumn(s):
    assert type(s) == pd.Series
    return pd.DataFrame(s)


class Location:

    """A point identified by (x,y) coordinates.

    supports: +, -, *, /, str, repr

    length  -- calculate length of vector to point from origin
    distance_to  -- calculate distance between two points
    as_tuple  -- construct tuple (x,y)
    clone  -- construct a duplicate
    integerize  -- convert x & y to integers
    floatize  -- convert x & y to floats
    move_to  -- reset x & y
    slide  -- move (in place) +dx, +dy, as spec'd by point
    slide_xy  -- move (in place) +dx, +dy
    rotate  -- rotate around the origin
    rotate_about  -- rotate around another point
    """

    #TODO: Validate arguments to __init__

    def __init__(self, x, y, sheet):
        self._setX(x)
        self._setY(y)
        self._setSheet(sheet)

    def __repr__(self):
        return '(%s, %s)' % (self.x, self.y)

    def asTuple(self):
        return (self.x, self.y)

    def _setX(self, x):
        self.x = x

    def _setY(self, y):
        self.y = y

    def _setSheet(self, s):
        assertType(s, Sheet)
        self.sheet = s

    def getSheet(self):
        return self.sheet

    def getX(self,):
        return self.x

    def getY(self):
        return self.y

    def __add__(self, p):
        """Point(x1+x2, y1+y2)"""
        assert self.getSheet() == p.getSheet()
        return Location(self.getX()+p.x, self.getY()+p.y, self.getSheet())

    def __sub__(self, p):
        """Point(x1-x2, y1-y2)"""
        assert self.getSheet() == p.getSheet()
        return Location(self.getX()-p.x, self.getY()-p.y, self.getSheet())

    def __mul__( self, scalar ):
        """Point(x1*x2, y1*y2)"""
        assert self.getSheet() == p.getSheet()
        return Location(self.getX()*scalar, self.getY()*scalar, self.getSheet())

    def __div__(self, scalar):
        """Point(x1/x2, y1/y2)"""
        assert self.getSheet() == p.getSheet()
        return Location(self.getX()/scalar, self.getY()/scalar, self.getSheet())

    def __repr__(self):
        return "(%r, %r)" % (self.getX(), self.getY())

    def copy(self):
        """Return a full copy of this point."""
        return Location(self.getX(), self.getY(), self.getSheet())

    def move_to(self, x, y):
        """Reset x & y coordinates."""
        self._setX(x)
        self._setY(y)

    def move_inplace(self, p):
        '''Move to new (x+dx,y+dy).

        Can anyone think up a better name for this function?
        slide? shift? delta? move_by?
        '''
        self._setX(self.getX() + p.x)
        self._setY(self.getY() + p.y)

    def move(self, p):
        c = self.copy()
        c._setX(c.getX() + p.x)
        c._setY(c.getY() + p.y)
        return c

    # def move_xy(self, dx, dy):
    #     '''Move to new (x+dx,y+dy).
    #
    #     Can anyone think up a better name for this function?
    #     slide? shift? delta? move_by?
    #     '''
    #     self._setX(self.getX() + dx)
    #     self._setY(self.getY() + dy)

    def toReference(self):
        return xl_rowcol_to_cell(self.getY(), self.getX())


class Cell():

    def __init__(self, location, data):
        assertType(location, Location)
        self._setData(data)
        self._setLocation(location)

    def __repr__(self):
        return 'Cell(' + self.toDataString() + ')'

    def _setData(self, data):
        self.data = data

    def getData(self):
        return self.data

    def _setLocation(self, location):
        self.location = location

    def getLocation(self):
        return self.location

    def copy(self):
        return Cell(self.getLocation(), self.getData())

    def move_inplace(self, p):
        self.getLocation().move_inplace(p)

    def getX(self):
        return self.getLocation().getX()

    def getY(self):
        return self.getLocation().getY()

    def toReferenceString(self, sheet):
        prepend = ("'%s'!" % self.getLocation().getSheet().getId()) if sheet != self.getLocation().getSheet() else ''
        return prepend + self.getLocation().toReference()

    def toDataString(self):
        if isinstance(self.getData(), Formula):
            return self.getData().toReferenceString(self.getLocation().getSheet())
        else:
            return str(self.getData())

    def toFinalString(self):
        if type(self.getData()) == Formula:
            prepend = '='
        else:
            prepend = ''
        return prepend + self.toDataString()



class Function():

    def __init__(self, string, position):
        self._setString(string)
        self._setPosition(position)

    def __repr__(self):
        return self.toString()

    def _setString(self, string):
        self.string = string

    @classmethod
    def subtract(cls):
        return Function('-', 'between')

    @classmethod
    def add(cls):
        return Function('+', 'between')

    @classmethod
    def divide(cls):
        return Function('/', 'between')

    @classmethod
    def multiply(cls):
        return Function('*', 'between')

    @classmethod
    def average(cls):
        return Function('AVERAGE', 'before')

    @classmethod
    def isEqual(cls):
        return Function('=', 'between')

    @classmethod
    def reference(cls):
        return Function('', 'between')

    @classmethod
    def range(cls):
        return Function(':', 'between')

    @classmethod
    def sum(cls):
        return Function('SUM', 'before')

    def toString(self):
        return self.string

    def toReferenceString(self, sheet=None):
        return self.string

    def _setPosition(self, position):
        acceptable_positions = ('before', 'between')
        assert position in acceptable_positions, 'Acceptable position arguments are %s.' % list(acceptable_positions)
        self.position = position

    def getPosition(self):
        return self.position

    def _mapFunctionToDFs(self, df1, df2, parentheses=False):
        """

        :param df1:
        :param df2:
        :param function:
        :param parentheses:
        :return:
        """

        assert df1.shape == df2.shape

        retval = pd.DataFrame(index=df1.index, columns=df1.columns)

        for col in range(len(df1.columns)):
            for row in range(len(df1)):
                retval[retval.columns[col]].iloc[row] = Formula(
                    self,
                    df1[df1.columns[col]].iloc[row],
                    df2[df2.columns[col]].iloc[row],
                    parentheses=parentheses)
        return retval

    def apply(self, *args, parentheses=False):
        """

        :param args:
        :param parentheses:
        :return:
        """

        def apply_helper(*args, joiner=None):
            args = list(args)
            for arg in range(len(args)):
                # This is contentious, but it's ambiguous if a Series should be a column or a row; if you wanna
                # use a row/column, slice it from the CellDF and then use asRow() or asColumn().
                assert type(args[arg]) != pd.Series
                if type(args[arg]) == Table:
                    args[arg] = args[arg].getCellDF()
            current = args[0]
            if len(args) == 1:
                return current
            else:
                # If dataframe is 1x1, just treat it as a cell since that is probably the expected behavior
                if isinstance(current, pd.DataFrame) and current.shape == (1, 1):
                    current = current[current.columns[0]].iloc[0]
                next = apply_helper(*args[1:], joiner=joiner)
                if isinstance(current, pd.DataFrame):
                    # Columns have shape[1]=1, rows have shape[0] = 1
                    if isinstance(next, pd.DataFrame):
                        if current.shape[0] == 1:
                            if next.shape[1] == 1:
                                return joiner._mapFunctionToDFs(next, pd.DataFrame(current).transpose())
                            else:
                                return joiner._mapFunctionToDFs(
                                    next,
                                    pd.concat([pd.DataFrame(current)]*next.shape[0], axis=0).reset_index(drop=True))
                        elif current.shape[1] == 1:
                            if next.shape[0] == 1:
                                return joiner._mapFunctionToDFs(next, current.transpose())
                            else:
                                current_formatted = pd.concat([pd.DataFrame(current)]*next.shape[1], axis=1)
                                current_formatted.columns = np.arange(len(current_formatted.columns))
                                return joiner._mapFunctionToDFs(next, current_formatted)
                        else:
                            if next.shape[0] == 1:
                                return joiner._mapFunctionToDFs(
                                    pd.concat([pd.DataFrame(next)]*current.shape[0], axis=0).reset_index(drop=True),
                                    current)
                            elif next.shape[1] == 1:
                                next_formatted = pd.concat([pd.DataFrame(next)]*current.shape[1], axis=1)
                                next_formatted.columns = np.arange(len(next_formatted.columns))
                                return joiner._mapFunctionToDFs(next_formatted, current)
                            else:
                                assert current.shape == next.shape,\
                                    'Operations on non-vector matrices of different shape are ambiguous'
                                return joiner._mapFunctionToDFs(next, current)
                    else:
                        next_formatted = pd.DataFrame(index=current.index, columns=current.columns).fillna(next)
                        return joiner._mapFunctionToDFs(next_formatted, current)
                else:
                    if isinstance(next, pd.DataFrame):
                        current_formatted = pd.DataFrame(index=next.index, columns=next.columns).fillna(current)
                        return joiner._mapFunctionToDFs(next, current_formatted)
                    else:
                        return Formula(joiner, next, current)


        if self.getPosition() == 'between':
            joiner = self
        else:
            joiner = Function(',', 'between')
        helper_out = apply_helper(*args[::-1], joiner=joiner)
        if isinstance(helper_out, Formula):
            helper_out._setParentheses(parentheses=parentheses)
        elif isinstance(helper_out, pd.DataFrame):
            for col in helper_out.columns:
                for row in helper_out.index:
                    if self.getPosition() == 'before':
                        helper_out[col].ix[row] = Formula(self, helper_out[col].ix[row], parentheses=parentheses)
                    else:
                        helper_out[col].ix[row]._setParentheses(parentheses=parentheses)
        return helper_out


class Formula():

    def __init__(self, function, *args, parentheses=False):
        # for arg in args:
        #     assert isinstance(arg, CellReference) or isinstance(arg, Formula)
        self._setFunction(function)
        self._setArgs(*args)
        self._setParentheses(parentheses)

    def __repr__(self):
        return self.toFinalString(None)

    @classmethod
    def range(cls, top_left, bottom_right, parentheses=False):
        return Function.range().apply(top_left, bottom_right, parentheses=parentheses)

    @classmethod
    def add(cls, *args, parentheses=False):
        return Function.add().apply(*args, parentheses=parentheses)

    @classmethod
    def subtract(cls, *args, parentheses=False):
        return Function.subtract().apply(*args, parentheses=parentheses)

    @classmethod
    def multiply(cls, *args, parentheses=False):
        return Function.multiply().apply(*args, parentheses=parentheses)

    @classmethod
    def divide(cls, *args, parentheses=False):
        return Function.divide().apply(*args, parentheses=parentheses)

    @classmethod
    def isEqual(cls, *args, parentheses=False):
        return Function.isEqual().apply(*args, parentheses=parentheses)

    @classmethod
    def sum(cls, *args, parentheses=False):
        return Function.sum().apply(*args, parentheses=parentheses)

    @classmethod
    def reference(cls, arg, parentheses=False):
        return Function.reference().apply(arg, parentheses=parentheses)

    # @classmethod
    # def reference(cls, *args, parentheses=False):
    #     return Function('', 'between').apply(*args, parentheses=parentheses)

    @classmethod
    def applyIf(cls, conditional, true, false, parentheses=False):
        if_func = Function('IF', 'before')
        return if_func.apply(conditional, true, false, parentheses=parentheses)

    def _setFunction(self, function):
        assertType(function, Function)
        self.function = function

    def getFunction(self):
        return self.function

    def _setArgs(self, *args):
        cleaned_args = []
        for arg in args:
            if isNumerical(arg):
                cleaned_args.append(str(arg))
            else:
                cleaned_args.append(arg)
        self.args = cleaned_args

    def getArgs(self):
        return self.args

    def _setParentheses(self, parentheses):
        self.parentheses = parentheses

    def getParentheses(self):
        return self.parentheses

    def toReferenceString(self, sheet=None):

        def toAppropriateString(arg):
            if isinstance(arg, (Cell, Function, Formula)):
                return arg.toReferenceString(sheet)
            else:
                return arg

        args = [toAppropriateString(arg) for arg in self.getArgs()]
        if self.getFunction().getPosition() == 'before':
            retval =  '%s(%s)' % (self.getFunction().toReferenceString(), ','.join(args))
        else:
            retval = '%s' % self.getFunction().toReferenceString().join(args)
        if self.getParentheses():
            return '(%s)' % retval
        else:
            return retval

    def toFinalString(self, sheet):
        return '=' + self.toReferenceString(sheet)


class Table():

    def __init__(
            self,
            id,
            data,
            location,
            total_row=False,
            include_header=True,
            include_index=True,
            include_id=True,
            body_style='general',
            **kwargs):
        self._setLocation(location)
        self._setSheet(location.getSheet())
        self._setId(id)
        self.total_row = total_row
        self.include_header = include_header
        self.include_index = include_index
        self.include_id = include_id
        self.body_style = body_style
        self._setDFs(data)
        self.data_width = self.getDataDF().shape[1]
        self.data_height = self.getDataDF().shape[0]


    def __repr__(self):
        return '%s\nAt location %s' % (self.getDataDF().__repr__(), self.getLocation().__repr__())

    def __getitem__(self, key):
        return self.getCellDF()[key]

    @property
    def ix(self):
        return self._ix

    @ix.setter
    def ix(self, index):
        self._ix = index

    @property
    def iloc(self):
        return self._iloc

    @iloc.setter
    def iloc(self, index):
        self._iloc = index

    def _setId(self, id):
        self.id = id

    def getId(self):
        return self.id

    def _setLocation(self, location):
        assertType(location, Location)
        self.location = location

    def getLocation(self):
        return self.location

    def _setSheet(self, s):
        assertType(s, Sheet)
        self.sheet = s

    def getSheet(self):
        return self.sheet

    def getDataDF(self):
        return self.data_df

    def _setDFs(self, to_df):
        if type(to_df) != pd.DataFrame:
            to_df = pd.DataFrame(to_df)
        self.data_df = to_df
        self.ix = self.getDataDF().ix
        self.iloc = self.getDataDF().iloc
        self._setCellDF()

    def getCellDF(self):
        return self.cell_df

    def _setCellDF(self):
        df = self.getDataDF()
        x_offset, y_offset = self.getDataOriginLocation().asTuple()
        d = df.to_dict()
        ordered_xs = df.columns
        ordered_ys = df.index
        c = OrderedDict()
        for x in range(len(ordered_xs)):
            xd = {}
            for y in range(len(ordered_ys)):
                 xd[ordered_ys[y]] = Cell(
                     Location(x+x_offset, y+y_offset, self.getSheet()),
                     d[ordered_xs[x]][ordered_ys[y]])
            c[ordered_xs[x]] = xd
        cell_df = pd.DataFrame(c, columns=ordered_xs, index=ordered_ys)
        if self.total_row:
            top = cell_df.iloc[0]
            bottom = cell_df.iloc[-1]
            total = Formula.sum(Formula.range(asRow(top), asRow(bottom)))
            for x in range(len(total.columns)):
                total[total.columns[x]].iloc[0] = Cell(bottom.iloc[x].getLocation().move(Location(0, 1, self.getSheet())), total[total.columns[x]].iloc[0])
            total.index = ['Total']
            cell_df = pd.concat([cell_df, total])
        self.cell_df = cell_df

    def getColumnHeaderHeight(self):
        return int(self.getIncludeHeader()) + int(self.getIncludeId())

    def getRowHeaderWidth(self):
        return int(self.getIncludeIndex())

    def getWidth(self):
        return self.data_width + self.getRowHeaderWidth()

    def getHeight(self):
        return self.data_height + self.getColumnHeaderHeight() + self.total_row

    def getDataOriginLocation(self):
        return self.getLocation() + Location(self.getRowHeaderWidth(), self.getColumnHeaderHeight(), self.getSheet())

    def getIncludeHeader(self):
        return self.include_header

    def getIncludeIndex(self):
        return self.include_index

    def getIncludeId(self):
        return self.include_id

    def getBodyStyle(self):
        return self.body_style

    def shift(self, dx, dy):
        # Note that this happens IN PLACE so that references to these cells will
        # be automatically adjusted when the table is exported to .xlsx
        l = Location(dx, dy, self.getSheet())
        self._setLocation(self.getLocation() + l)
        self.getCellDF().applymap(lambda cell: cell.move_inplace(Location(dx, dy, self.getSheet())))

    def shiftToLocation(self, end):
        assert type(end) == Location
        start = self.getLocation()
        diff = end - start
        self.shift(diff.asTuple()[0], diff.asTuple()[1])

    def getDataRangeReference(self, parentheses=False):
        df = self.getCellDF()
        return Formula.range(df[df.columns[0]].iloc[0], df[df.columns[-1]].iloc[-1])

    def toRetentionRate(self):
        ret = self.getCellDF().copy()
        for row in range(len(ret)):
            for col in range(row, len(ret.columns)):
                num = ret[ret.columns[col]].iloc[row]
                div = ret[ret.columns[row]].iloc[row]
                ret[ret.columns[col]].iloc[row] = Formula.divide(num, div)
        return ret

    def toAnnualRenewalWaterfall(self, renewal_rates):
        renewal_rates = renewal_rates.getCellDF()
        adds = self.getCellDF()
        ret_all = {}
        for col in adds:
            this_adds = adds[col]
            this_rr = renewal_rates[col]
            all_dates = this_adds.index
            retval = pd.DataFrame(index=all_dates, columns=all_dates).to_dict()
            for cohort in all_dates:
                cohort_dict = {}
                annuals = pd.date_range(cohort, max(all_dates), freq='12MS')
                base = this_adds.ix[cohort]
                cohort_dict[annuals[0]] = Formula(Function.reference(), base)
                for year in range(1, len(annuals)):
                    cohort_dict[annuals[year]] = Formula(Function.multiply(), base, this_rr.ix[year])
                retval[cohort] = cohort_dict
            ret_all[col] = pd.DataFrame.from_dict(retval, orient='index')
        return ret_all

    def periodsFromForecastDate(self, num_periods_out, as_data=True):
        cdf = self.getCellDF().copy()
        ddf = cdf.applymap(lambda c: c.getData())
        assert ('CountDate' in cdf.columns and 'ForecastDate' in cdf.columns and 'Count'  in cdf.columns)
        ddf = ddf.set_index('ForecastDate')
        cdf.index = ddf.index
        to_df = {}
        for i in cdf.index.unique():
            ix = ddf.ix[i].set_index('CountDate').index
            this_index = cdf.ix[i]
            this_index.index = ix
            first_ind = np.where(this_index.index == i)[0][0]
            assert this_index.index.is_monotonic
            # assert len(this_index['m'].unique() == 1)
            # m = this_index['m'].unique()[0]
            # c = this_index['Count'].iloc[first_ind]
            this_index = this_index['Count'].iloc[first_ind+1:first_ind+1+num_periods_out]
            this_index.index = np.arange(1, num_periods_out+1)
            # this_index.ix['m'] = m
            # this_index.ix['c'] = c
            to_df[i] = this_index

        if as_data:
            return pd.DataFrame(to_df)
        else:
            return pd.DataFrame(to_df).applymap(lambda c: Formula(Function.reference(), c))


class Sheet():

    def __init__(self, id, wb):
        self._setId(id)
        self._setWorkbook(wb)
        self.tables = OrderedDict()
        self.next_table_origin = Location(0, 0, self)

    def __repr__(self):
        return '\n\n'.join([table.__repr__() for table in self.getTables().values()])

    def _setId(self, id):
        self.id = id

    def _setWorkbook(self, wb):
        assertType(wb, Workbook)
        self.workbook = wb

    def getWorkbook(self):
        return self.workbook

    def getId(self):
        return self.id

    def getNextTableOrigin(self, relative_position='below', margin=1):
        if len(self.getTables()) == 0:
            return Location(0, 0, self)
        else:
            last = list(self.getTables().items())[-1][1]
            if relative_position == 'right':
                return Location(last.getLocation().getX() + last.getWidth() + margin, last.getLocation().getY(), self)
            else:
                return Location(0, last.getLocation().getY() + last.getHeight() + margin, self)

    def addTable(self, id, df, relative_position='below', total_row=False, margin=1, include_header=True, include_index=True, include_id=True, body_style='general'):
        location = self.getNextTableOrigin(relative_position=relative_position, margin=margin)
        table = Table(id, df, location, total_row=total_row, include_header=include_header, include_index=include_index, include_id=include_id, body_style=body_style)
        self.tables[table.getId()] = table
        return table

    def getTables(self):
        return self.tables

    def getTable(self, id):
        return self.getTables()[id]

class Workbook():

    def __init__(self, id):
        self._setId(id)
        self.sheets = OrderedDict()

    def __repr__(self):
        return '\n----------\n\n'.join(['%s:\n\n%s' % (sheet.getId(),sheet.__repr__()) for sheet in self.getSheets().values()])

    def addSheet(self, id):
        sheet = Sheet(id, self)
        sheet._setWorkbook(self)
        self.sheets[sheet.getId()] = sheet
        return sheet

    def getSheets(self):
        return self.sheets

    def _setId(self, id):
        self.id = id

    def getId(self):
        return self.id

    def _getStylesDict(self, workbook):
        styles = {}
        styles['general'] = workbook.add_format({
            'font_name': 'Arial',
            'num_format': '###,###,###'
        })
        styles['general_date'] = workbook.add_format({
            'font_name': 'Arial',
            'num_format': 'mm/dd/yy'
        })
        styles['bold'] = workbook.add_format({
            'bold': True,
            'font_name': 'Arial'})
        styles['money'] = workbook.add_format({
            'font_name': 'Arial',
            'num_format': '$###,###,###'
        })
        styles['column_header'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'bg_color': '#404040'
        })
        styles['column_header_date'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'num_format': 'mm/dd/yy',
            'bg_color': '#404040'
        })
        styles['row_header'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'bg_color': '#A6A6A6'
        })
        styles['row_header_date'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'num_format': 'mm/dd/yy',
            'bg_color': '#A6A6A6'
        })
        styles['columns_title'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'bg_color': '#404040',
            'bold': True
        })
        styles['columns_title_date'] = workbook.add_format({
            'font_color': '#FFFFFF',
            'font_name': 'Arial',
            'bg_color': '#404040',
            'num_format': 'mm/dd/yy',
            'bold': True
        })
        return styles

    def _insertSheet(self, sheet):
        sheet._setWorkbook(self)
        self.sheets[sheet.getId()] = sheet
        return sheet

    def exportAsXLSX(self, path):
        dir = '/'.join(path.split('/')[:-1]) + '/'
        if not os.path.exists(dir):
            os.makedirs(dir)
        workbook = xlsxwriter.Workbook(path, {'default_date_format': 'mm/dd/yy'})
        styles = self._getStylesDict(workbook)
        for sheet in self.getSheets().values():
            xlsx_sheet = workbook.add_worksheet(name=sheet.getId())
            def writeFunction(y, x, d, style_prefix):
                if isDatetimeLike(d):
                    style_suffix = '_date'
                    if type(d) == dt.date:
                        xlsx_sheet.write_datetime(y, x, d, styles[style_prefix + style_suffix])
                    else:
                        xlsx_sheet.write_datetime(y, x, d.date(), styles[style_prefix + style_suffix])
                elif isinstance(d, Formula):
                    xlsx_sheet.write(y, x, d.toFinalString(sheet), styles[style_prefix])
                else:
                    try:
                        xlsx_sheet.write(y, x, d, styles[style_prefix])
                    except TypeError:
                        # This is specifically intended to handle NaNs, so that they aren't written at all
                        pass
            for table in sheet.getTables().values():
                x_offset, y_offset = table.getLocation().asTuple()
                row_counter = 0
                if table.getIncludeId():
                    writeFunction(y_offset, x_offset, table.getId(), 'columns_title')
                    for col in range(1, table.getWidth()):
                        writeFunction(y_offset, x_offset+col, '', 'columns_title')
                    row_counter += 1
                if table.getIncludeHeader():
                    col_counter = 0
                    if table.getIncludeIndex():
                        ind_offset = 1
                        writeFunction(y_offset+row_counter, x_offset+col_counter, '', 'column_header')
                    else:
                        ind_offset=0
                    for column in table.getCellDF().columns:
                        writeFunction(y_offset+row_counter, x_offset+ind_offset+col_counter, column, 'column_header')
                        col_counter += 1
                    row_counter += 1
                for row in table.getCellDF().iterrows():
                    col_counter = 0
                    if table.getIncludeIndex():
                        writeFunction(row_counter+y_offset, x_offset, row[0], 'row_header')
                        col_counter += 1
                    for i in row[1].index:
                        d = row[1][i].getData()
                        writeFunction(row_counter+y_offset, col_counter+x_offset, d, table.getBodyStyle())
                        col_counter += 1
                    row_counter += 1
        workbook.close()

    def exportAsXLSXandCSVs(self, path):
        self.exportAsXLSX(path)
        wb = openpyxl.load_workbook(path)
        dir = '/'.join(path.split('/')[:-1]) + '/'
        for i in self.getSheets():
            sh = wb.get_sheet_by_name(i)
            this_csv = open(dir + i + '.csv', 'w')
            wr = csv.writer(this_csv, quoting=csv.QUOTE_ALL)
            for row in sh.rows:
                wr.writerow([x.value for x in row])
            this_csv.close()